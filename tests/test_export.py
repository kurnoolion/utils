from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from disk_cleaner.export import HEADERS, human_size, parse_size, run_export
from disk_cleaner.finalize import run_finalize
from disk_cleaner.scan import run_scan


# --- pure helpers --------------------------------------------------------

def test_parse_size_units():
    assert parse_size("0") == 0
    assert parse_size("1024") == 1024            # bare = bytes
    assert parse_size("1KB") == 1024
    assert parse_size("100MB") == 100 * 1024**2
    assert parse_size("1.5 GB") == int(1.5 * 1024**3)
    assert parse_size("2 TB") == 2 * 1024**4
    assert parse_size("  100 mb ") == 100 * 1024**2


def test_human_size_formats():
    assert human_size(None) == ""
    assert human_size(0) == "0.0 B"
    assert human_size(1024) == "1.0 KB"
    assert human_size(int(1.5 * 1024**2)) == "1.5 MB"
    assert human_size(2 * 1024**3) == "2.0 GB"


# --- end-to-end export ---------------------------------------------------

def _build_tree(root: Path) -> None:
    """Sizes chosen so we can test threshold filtering precisely.
        root/
          tiny.txt           (10 bytes)        — below file threshold
          medium.bin         (5 KB)            — above 1 KB threshold
          big.bin            (50 KB)           — top
          subA/              (folder size: 5 KB)
            inside.bin       (5 KB)
          subB/              (folder size: 0)  — no files
    """
    (root / "tiny.txt").write_bytes(b"x" * 10)
    (root / "medium.bin").write_bytes(b"m" * 5 * 1024)
    (root / "big.bin").write_bytes(b"B" * 50 * 1024)
    (root / "subA").mkdir()
    (root / "subA" / "inside.bin").write_bytes(b"i" * 5 * 1024)
    (root / "subB").mkdir()


def _setup(tmp_path):
    root = tmp_path / "tree"
    root.mkdir()
    _build_tree(root)
    db = tmp_path / "scan.db"
    run_scan(str(root), str(db), resume=False)
    run_finalize(str(db))
    return root, db


def test_export_produces_xlsx_with_expected_headers(tmp_path):
    _, db = _setup(tmp_path)
    out = tmp_path / "review.xlsx"
    run_export(str(db), str(out), "1KB", "1KB")

    wb = load_workbook(str(out))
    ws = wb["Review"]
    headers = [c.value for c in ws[1]]
    assert headers == HEADERS


def test_export_filters_by_size_thresholds(tmp_path):
    _, db = _setup(tmp_path)
    out = tmp_path / "review.xlsx"
    # 1 KB file threshold excludes tiny.txt; 100 KB folder threshold excludes both folders.
    run_export(str(db), str(out), "1KB", "100KB")

    wb = load_workbook(str(out))
    ws = wb["Review"]
    paths = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]

    assert "tiny.txt" not in paths           # too small
    assert "medium.bin" in paths
    assert "big.bin" in paths
    assert "subA" not in paths               # 5 KB < 100 KB threshold
    assert "subB" not in paths


def test_export_includes_folders_above_folder_threshold(tmp_path):
    _, db = _setup(tmp_path)
    out = tmp_path / "review.xlsx"
    run_export(str(db), str(out), "1KB", "1KB")

    wb = load_workbook(str(out))
    ws = wb["Review"]
    rows = {ws.cell(r, 1).value: ws.cell(r, 4).value for r in range(2, ws.max_row + 1)}
    assert rows["subA"] == "Folder"
    assert "subB" not in rows                # 0 < 1KB


def test_export_sizes_in_kb_and_human(tmp_path):
    _, db = _setup(tmp_path)
    out = tmp_path / "review.xlsx"
    run_export(str(db), str(out), "1KB", "100GB")  # exclude folders

    wb = load_workbook(str(out))
    ws = wb["Review"]
    by_path = {
        ws.cell(r, 1).value: (ws.cell(r, 2).value, ws.cell(r, 3).value)
        for r in range(2, ws.max_row + 1)
    }
    # big.bin is 50 KB exactly
    assert by_path["big.bin"][0] == 50.0
    assert by_path["big.bin"][1] == "50.0 KB"
    # medium.bin is 5 KB exactly
    assert by_path["medium.bin"][0] == 5.0


def test_export_delete_column_defaults_to_N(tmp_path):
    _, db = _setup(tmp_path)
    out = tmp_path / "review.xlsx"
    run_export(str(db), str(out), "1KB", "1KB")

    wb = load_workbook(str(out))
    ws = wb["Review"]
    for r in range(2, ws.max_row + 1):
        assert ws.cell(r, 10).value == "N"


def test_export_sorted_by_size_descending(tmp_path):
    _, db = _setup(tmp_path)
    out = tmp_path / "review.xlsx"
    run_export(str(db), str(out), "1KB", "100GB")

    wb = load_workbook(str(out))
    ws = wb["Review"]
    sizes_kb = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1)]
    assert sizes_kb == sorted(sizes_kb, reverse=True)


def test_export_dates_are_datetimes(tmp_path):
    _, db = _setup(tmp_path)
    out = tmp_path / "review.xlsx"
    run_export(str(db), str(out), "1KB", "1KB")

    wb = load_workbook(str(out))
    ws = wb["Review"]
    for r in range(2, ws.max_row + 1):
        for col in (7, 8, 9):  # Created, Modified, Accessed
            value = ws.cell(r, col).value
            assert isinstance(value, datetime), (
                f"row {r} col {col} expected datetime, got {type(value)}"
            )


def test_export_meta_sheet_present(tmp_path):
    _, db = _setup(tmp_path)
    out = tmp_path / "review.xlsx"
    run_export(str(db), str(out), "1KB", "1KB")

    wb = load_workbook(str(out))
    assert "Meta" in wb.sheetnames
    ws = wb["Meta"]
    keys = {ws.cell(r, 1).value for r in range(1, ws.max_row + 1)}
    assert "scan_root_original" in keys
    assert "finalized_at" in keys
    assert "export_min_file_size_bytes" in keys
    assert "export_row_count" in keys


def test_export_refuses_unfinalized_db(tmp_path):
    root = tmp_path / "tree"
    root.mkdir()
    _build_tree(root)
    db = tmp_path / "scan.db"
    run_scan(str(root), str(db), resume=False)
    # NOTE: no finalize call

    out = tmp_path / "review.xlsx"
    with pytest.raises(SystemExit, match="not been finalized"):
        run_export(str(db), str(out), "1KB", "1KB")


def test_export_refuses_db_without_scan(tmp_path):
    import sqlite3
    db = tmp_path / "empty.db"
    sqlite3.connect(str(db)).close()
    out = tmp_path / "review.xlsx"
    with pytest.raises(SystemExit, match="no scan"):
        run_export(str(db), str(out), "1KB", "1KB")
