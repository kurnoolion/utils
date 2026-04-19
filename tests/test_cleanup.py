import os
import sqlite3
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from disk_cleaner.cleanup import (
    REVIEW_REQUIRED_COLUMNS,
    _filter_redundant_descendants,
    _same_volume,
    run_cleanup,
)
from disk_cleaner.export import HEADERS, run_export
from disk_cleaner.finalize import run_finalize
from disk_cleaner.scan import run_scan


# ----- shared helpers ----------------------------------------------------

def _build_tree(root: Path) -> None:
    """root/
         keep.txt           (10 bytes)
         delete_me.bin      (5 KB)
         folderA/
           a1.txt           (1 KB)
           a2.txt           (1 KB)
         folderB/
           inner/
             deep.bin       (3 KB)
    """
    (root / "keep.txt").write_bytes(b"x" * 10)
    (root / "delete_me.bin").write_bytes(b"d" * 5 * 1024)
    (root / "folderA").mkdir()
    (root / "folderA" / "a1.txt").write_bytes(b"a" * 1024)
    (root / "folderA" / "a2.txt").write_bytes(b"b" * 1024)
    (root / "folderB" / "inner").mkdir(parents=True)
    (root / "folderB" / "inner" / "deep.bin").write_bytes(b"z" * 3 * 1024)


def _make_review(out: Path, marks: dict[str, str]) -> None:
    """Write a minimal review.xlsx. `marks` maps rel_path -> 'Y' or 'N'.

    Each marked path's row gets Type=Folder if path looks like a directory
    (no extension) and File otherwise. That's a heuristic for the test;
    the cleanup code only uses Type to label moves, not to drive behavior.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.append(HEADERS)
    for path, flag in marks.items():
        kind = "Folder" if "." not in os.path.basename(path) else "File"
        # Columns: Path, Size(KB), Size(Human), Type, Category, Extension,
        #          Created, Modified, Accessed, Delete?
        ws.append([path, None, None, kind, None, None, None, None, None, flag])
    wb.save(str(out))


def _setup_full(tmp_path: Path):
    """Run scan + finalize so the DB is ready for cleanup tests."""
    root = tmp_path / "src"
    root.mkdir()
    _build_tree(root)
    db = tmp_path / "scan.db"
    run_scan(str(root), str(db), resume=False)
    run_finalize(str(db))
    trash = tmp_path / "trash"
    trash.mkdir()
    log_dir = tmp_path / "logs"
    return root, db, trash, log_dir


def _moves(db: Path) -> dict[str, tuple]:
    conn = sqlite3.connect(str(db))
    try:
        return {
            row[0]: row[1:]
            for row in conn.execute(
                "SELECT rel_path, kind, status, error, dest_path FROM moves"
            )
        }
    finally:
        conn.close()


# ----- pure-function tests -----------------------------------------------

def test_filter_redundant_descendants_drops_children_when_parent_marked():
    marked = [
        ("folderA", "folder"),
        ("folderA/a1.txt", "file"),
        ("folderA/a2.txt", "file"),
        ("loose.txt", "file"),
    ]
    assert _filter_redundant_descendants(marked) == [
        ("folderA", "folder"),
        ("loose.txt", "file"),
    ]


def test_filter_redundant_descendants_handles_deep_nesting():
    marked = [
        ("a", "folder"),
        ("a/b/c/d.txt", "file"),
        ("a/b", "folder"),
        ("x.txt", "file"),
    ]
    keep = _filter_redundant_descendants(marked)
    assert ("a", "folder") in keep
    assert ("x.txt", "file") in keep
    assert ("a/b/c/d.txt", "file") not in keep
    assert ("a/b", "folder") not in keep


def test_same_volume_true_for_subpath_of_root(tmp_path):
    sub = tmp_path / "sub"
    sub.mkdir()
    assert _same_volume(str(tmp_path), str(sub)) is True


# ----- end-to-end cleanup tests ------------------------------------------

def test_cleanup_moves_marked_file(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"delete_me.bin": "Y", "keep.txt": "N"})

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    assert not (root / "delete_me.bin").exists()
    assert (trash / "delete_me.bin").exists()
    assert (root / "keep.txt").exists()                 # untouched

    moves = _moves(db)
    assert moves["delete_me.bin"][1] == "moved"


def test_cleanup_moves_marked_folder_recursively(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"folderA": "Y"})

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    assert not (root / "folderA").exists()
    assert (trash / "folderA" / "a1.txt").exists()
    assert (trash / "folderA" / "a2.txt").exists()


def test_cleanup_skips_descendants_when_ancestor_also_marked(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(
        review,
        {
            "folderA": "Y",
            "folderA/a1.txt": "Y",   # redundant — should be dropped from plan
            "folderA/a2.txt": "Y",   # redundant
        },
    )

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    moves = _moves(db)
    # Only the folder gets recorded as moved — children never attempted
    assert "folderA" in moves
    assert "folderA/a1.txt" not in moves
    assert "folderA/a2.txt" not in moves
    # Folder structure preserved in trash
    assert (trash / "folderA" / "a1.txt").exists()


def test_cleanup_y_is_case_insensitive(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"delete_me.bin": "y"})  # lowercase

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))
    assert (trash / "delete_me.bin").exists()


def test_cleanup_ignores_unmarked_and_blank_rows(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"keep.txt": "N", "delete_me.bin": ""})

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    assert (root / "keep.txt").exists()
    assert (root / "delete_me.bin").exists()
    assert _moves(db) == {}


def test_cleanup_handles_missing_source(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"never_existed.bin": "Y"})

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    moves = _moves(db)
    assert moves["never_existed.bin"][1] == "source_missing"


def test_cleanup_records_failures_when_dest_already_exists(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    # Pre-create the dest path so the move collides
    (trash / "delete_me.bin").write_bytes(b"prior")

    review = tmp_path / "review.xlsx"
    _make_review(review, {"delete_me.bin": "Y"})

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    moves = _moves(db)
    assert moves["delete_me.bin"][1] == "failed"
    assert "already exists" in moves["delete_me.bin"][2]
    # Source still in place — not partially moved
    assert (root / "delete_me.bin").exists()


def test_cleanup_resume_skips_completed_and_retries_failed(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"delete_me.bin": "Y", "folderA": "Y"})

    # First run: pre-fill the dest of folderA so it fails
    (trash / "folderA").mkdir()
    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    moves = _moves(db)
    assert moves["delete_me.bin"][1] == "moved"
    assert moves["folderA"][1] == "failed"

    # Fix the obstacle and resume
    (trash / "folderA").rmdir()
    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    moves = _moves(db)
    assert moves["folderA"][1] == "moved"          # retried successfully
    assert moves["delete_me.bin"][1] == "moved"     # not re-attempted (still moved)


def test_cleanup_dry_run_does_not_move_or_record(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"delete_me.bin": "Y", "folderA": "Y"})

    run_cleanup(
        str(db), str(review), str(root), str(trash), str(log_dir),
        dry_run=True,
    )

    assert (root / "delete_me.bin").exists()
    assert (root / "folderA").exists()
    assert not list(trash.iterdir())
    assert _moves(db) == {}


def test_cleanup_writes_log_files(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"delete_me.bin": "Y", "missing.bin": "Y"})

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    main_logs = list(log_dir.glob("cleanup_*.log"))
    main_logs = [p for p in main_logs if "_errors" not in p.name]
    err_logs = list(log_dir.glob("cleanup_*_errors.log"))

    assert len(main_logs) == 1
    assert len(err_logs) == 1

    main_text = main_logs[0].read_text()
    err_text = err_logs[0].read_text()

    assert "cleanup start" in main_text
    assert "OK delete_me.bin" in main_text
    assert "MISSING missing.bin" in main_text

    # Errors-only log should NOT have the OK line, but SHOULD have the missing one
    assert "OK delete_me.bin" not in err_text
    assert "MISSING missing.bin" in err_text


def test_cleanup_validates_required_columns(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)

    bad_review = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.append(["Path", "Size (KB)"])  # missing Type and Delete?
    ws.append(["delete_me.bin", 5])
    wb.save(str(bad_review))

    with pytest.raises(SystemExit, match="missing required columns"):
        run_cleanup(str(db), str(bad_review), str(root), str(trash), str(log_dir))


def test_cleanup_refuses_missing_root(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"delete_me.bin": "Y"})

    with pytest.raises(SystemExit, match="not found"):
        run_cleanup(str(db), str(review), str(tmp_path / "no_such_root"),
                    str(trash), str(log_dir))


def test_cleanup_refuses_missing_review(tmp_path):
    root, db, trash, log_dir = _setup_full(tmp_path)

    with pytest.raises(SystemExit, match="--review not found"):
        run_cleanup(str(db), str(tmp_path / "no.xlsx"), str(root),
                    str(trash), str(log_dir))


def test_required_columns_constant_unchanged():
    # Guard against accidental change of the required column set.
    assert REVIEW_REQUIRED_COLUMNS == {"Path", "Type", "Delete?"}


@pytest.mark.skipif(sys.platform == "win32", reason="cross-volume check uses st_dev on POSIX")
def test_cleanup_refuses_cross_volume_by_default(tmp_path, monkeypatch):
    """Force _same_volume to report False and ensure cleanup refuses without flag."""
    from disk_cleaner import cleanup as cleanup_mod

    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    _make_review(review, {"delete_me.bin": "Y"})

    monkeypatch.setattr(cleanup_mod, "_same_volume", lambda a, b: False)

    with pytest.raises(SystemExit, match="not on the same volume"):
        run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    # And succeeds when override flag is set
    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir),
                allow_cross_volume=True)
    assert (trash / "delete_me.bin").exists()


def test_cleanup_round_trip_uses_real_review_export(tmp_path):
    """Exercise the full pipeline: scan -> finalize -> export -> mark Y -> cleanup."""
    root, db, trash, log_dir = _setup_full(tmp_path)
    review = tmp_path / "review.xlsx"
    run_export(str(db), str(review), "1KB", "1KB")

    # Mark folderA Y in the exported xlsx and save
    wb = load_workbook(str(review))
    ws = wb["Review"]
    header = [c.value for c in ws[1]]
    col_path = header.index("Path") + 1
    col_delete = header.index("Delete?") + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col_path).value == "folderA":
            ws.cell(r, col_delete).value = "Y"
    wb.save(str(review))

    run_cleanup(str(db), str(review), str(root), str(trash), str(log_dir))

    assert not (root / "folderA").exists()
    assert (trash / "folderA" / "a1.txt").exists()
