import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .db import connect, get_meta, has_scan

log = logging.getLogger(__name__)

EXCEL_MAX_ROWS = 1_048_576

HEADERS = [
    "Path",
    "Size (KB)",
    "Size (Human)",
    "Type",
    "Category",
    "Extension",
    "Created",
    "Modified",
    "Accessed",
    "Delete?",
]

_UNITS = {"B": 1, "KB": 1024, "MB": 1024**2, "GB": 1024**3, "TB": 1024**4}


def parse_size(s: str) -> int:
    """Parse e.g. '100MB', '1.5 GB', or a bare byte count."""
    s = s.strip().upper().replace(" ", "")
    for unit in sorted(_UNITS, key=len, reverse=True):
        if s.endswith(unit):
            return int(float(s[: -len(unit)]) * _UNITS[unit])
    return int(s)


def human_size(n: int | None) -> str:
    if n is None:
        return ""
    v = float(n)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if abs(v) < 1024:
            return f"{v:.1f} {unit}"
        v /= 1024
    return f"{v:.1f} PB"


def _to_dt(ts):
    if ts is None:
        return None
    try:
        return datetime.fromtimestamp(ts)
    except (OSError, ValueError, OverflowError):
        return None


def run_export(
    db_path: str,
    out_path: str,
    min_file_size_str: str,
    min_folder_size_str: str,
) -> None:
    min_file = parse_size(min_file_size_str)
    min_folder = parse_size(min_folder_size_str)

    conn = connect(db_path)
    try:
        if not has_scan(conn):
            raise SystemExit(f"DB has no scan: {db_path}")
        if get_meta(conn, "finalized_at") is None:
            raise SystemExit(
                "DB has not been finalized. Run "
                "`python -m disk_cleaner finalize --db ...` first."
            )

        rows = conn.execute(
            "SELECT path, kind, size_bytes, extension, category, "
            "       ctime, mtime, atime "
            "FROM entries "
            "WHERE (kind='file'   AND size_bytes >= ?) "
            "   OR (kind='folder' AND size_bytes >= ?) "
            "ORDER BY size_bytes DESC, path ASC",
            (min_file, min_folder),
        ).fetchall()

        log.info(
            "Selected %d rows (file>=%s, folder>=%s)",
            len(rows),
            human_size(min_file),
            human_size(min_folder),
        )

        if len(rows) + 1 > EXCEL_MAX_ROWS:
            raise SystemExit(
                f"Too many rows for Excel ({len(rows)} > {EXCEL_MAX_ROWS - 1}). "
                f"Raise --min-file-size or --min-folder-size."
            )

        wb = Workbook()
        _write_review_sheet(wb, rows)
        _write_meta_sheet(wb, conn, min_file, min_folder, len(rows))
        wb.save(out_path)
        log.info("Wrote %s", out_path)
    finally:
        conn.close()


def _write_review_sheet(wb: Workbook, rows: list[tuple]) -> None:
    ws = wb.active
    ws.title = "Review"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for path, kind, size_b, ext, cat, ctime, mtime, atime in rows:
        size_kb = round(size_b / 1024, 2) if size_b is not None else None
        ws.append([
            path,
            size_kb,
            human_size(size_b),
            "Folder" if kind == "folder" else "File",
            cat or "",
            ext or "",
            _to_dt(ctime),
            _to_dt(mtime),
            _to_dt(atime),
            "N",
        ])

    last_row = ws.max_row
    if last_row >= 2:
        for col in (7, 8, 9):  # Created, Modified, Accessed
            for r in range(2, last_row + 1):
                ws.cell(r, col).number_format = "yyyy-mm-dd hh:mm:ss"

        dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
        dv.add(f"J2:J{last_row}")
        ws.add_data_validation(dv)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"

    widths = {1: 70, 2: 14, 3: 14, 4: 8, 5: 10, 6: 12, 7: 19, 8: 19, 9: 19, 10: 9}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_meta_sheet(wb: Workbook, conn, min_file: int, min_folder: int, n_rows: int) -> None:
    ws = wb.create_sheet("Meta")
    ws.append(["Key", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for key, value in conn.execute(
        "SELECT key, value FROM scan_meta ORDER BY key"
    ):
        ws.append([key, value])

    ws.append(["export_min_file_size_bytes", str(min_file)])
    ws.append(["export_min_folder_size_bytes", str(min_folder)])
    ws.append(["export_row_count", str(n_rows)])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
