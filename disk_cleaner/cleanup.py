import logging
import os
import shutil
import time
from datetime import datetime

from openpyxl import load_workbook

from .db import connect, init_schema

log = logging.getLogger(__name__)

REVIEW_REQUIRED_COLUMNS = {"Path", "Type", "Delete?"}

STATUS_MOVED = "moved"
STATUS_MISSING = "source_missing"
STATUS_FAILED = "failed"
TERMINAL_STATUSES = (STATUS_MOVED, STATUS_MISSING)


def run_cleanup(
    db_path: str,
    review_path: str,
    root: str,
    trash: str,
    log_dir: str,
    allow_cross_volume: bool = False,
    dry_run: bool = False,
) -> None:
    root_abs = os.path.abspath(root)
    trash_abs = os.path.abspath(trash)
    log_dir_abs = os.path.abspath(log_dir)

    if not os.path.isdir(root_abs):
        raise SystemExit(f"--root not found or not a directory: {root_abs}")
    if not os.path.isfile(review_path):
        raise SystemExit(f"--review not found: {review_path}")

    os.makedirs(log_dir_abs, exist_ok=True)
    handlers, prev_level = _setup_file_log(log_dir_abs)

    try:
        log.info("=== cleanup start (dry_run=%s) ===", dry_run)
        log.info("root=%s", root_abs)
        log.info("trash=%s", trash_abs)
        log.info("review=%s", review_path)

        if not _same_volume(root_abs, trash_abs) and not allow_cross_volume:
            raise SystemExit(
                f"Trash '{trash_abs}' is not on the same volume as root "
                f"'{root_abs}'. Pass --allow-cross-volume to override "
                f"(slow, doubles I/O, can fill destination)."
            )

        marked = _read_marked_rows(review_path)
        log.info("Marked rows in review: %d", len(marked))

        plan = _filter_redundant_descendants(marked)
        skipped_overlap = len(marked) - len(plan)
        if skipped_overlap:
            log.info("Dropped %d marked entries whose ancestor is also marked", skipped_overlap)

        # Process largest first so we make visible progress on space
        # (folders before contained files when sizes happen to match).
        conn = connect(db_path)
        try:
            init_schema(conn)
            done = _load_terminal_moves(conn)
            log.info("Resume: %d entries already in terminal state", len(done))

            counts = {"moved": 0, "missing": 0, "failed": 0, "skipped_done": 0}

            for rel_path, kind in plan:
                if rel_path in done:
                    counts["skipped_done"] += 1
                    continue

                src = os.path.join(root_abs, _to_native(rel_path))
                dst = os.path.join(trash_abs, _to_native(rel_path))

                if not os.path.lexists(src):
                    log.warning("MISSING %s (source not found)", rel_path)
                    counts["missing"] += 1
                    if not dry_run:
                        _record_move(conn, rel_path, kind, None, dst,
                                     STATUS_MISSING, "source not found at cleanup time")
                    continue

                size_bytes = _safe_size(src)

                if dry_run:
                    log.info("DRY-RUN would move %s -> %s (%s bytes)",
                             rel_path, dst, size_bytes)
                    continue

                try:
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    if os.path.lexists(dst):
                        raise FileExistsError(f"destination already exists: {dst}")
                    shutil.move(src, dst)
                except Exception as e:
                    log.error("FAIL %s: %s", rel_path, e)
                    counts["failed"] += 1
                    _record_move(conn, rel_path, kind, size_bytes, dst,
                                 STATUS_FAILED, str(e))
                    continue

                log.info("OK %s -> %s", rel_path, dst)
                counts["moved"] += 1
                _record_move(conn, rel_path, kind, size_bytes, dst,
                             STATUS_MOVED, None)

            log.info(
                "Cleanup summary: moved=%d missing=%d failed=%d skipped_done=%d",
                counts["moved"], counts["missing"], counts["failed"], counts["skipped_done"],
            )
        finally:
            conn.close()
    finally:
        _teardown_file_log(handlers, prev_level)


def _setup_file_log(log_dir: str):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    main_path = os.path.join(log_dir, f"cleanup_{timestamp}.log")
    main_h = logging.FileHandler(main_path)
    main_h.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    main_h.setLevel(logging.INFO)

    err_path = os.path.join(log_dir, f"cleanup_{timestamp}_errors.log")
    err_h = logging.FileHandler(err_path)
    err_h.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    err_h.setLevel(logging.WARNING)  # WARNING (missing) and ERROR (failures)

    root = logging.getLogger()
    root.addHandler(main_h)
    root.addHandler(err_h)
    prev_level = root.level
    if root.level > logging.INFO or root.level == 0:
        root.setLevel(logging.INFO)
    return [main_h, err_h], prev_level


def _teardown_file_log(handlers, prev_level):
    root = logging.getLogger()
    for h in handlers:
        root.removeHandler(h)
        h.close()
    root.setLevel(prev_level)


def _same_volume(a: str, b: str) -> bool:
    if os.name == "nt":
        a_drive = os.path.splitdrive(os.path.abspath(a))[0].lower()
        b_drive = os.path.splitdrive(os.path.abspath(b))[0].lower()
        return a_drive == b_drive

    a_dev = os.stat(a).st_dev
    p = os.path.abspath(b)
    while not os.path.exists(p):
        parent = os.path.dirname(p)
        if parent == p:
            return False
        p = parent
    return os.stat(p).st_dev == a_dev


def _read_marked_rows(review_path: str) -> list[tuple[str, str]]:
    wb = load_workbook(review_path, read_only=True, data_only=True)
    try:
        if "Review" not in wb.sheetnames:
            raise SystemExit(f"Review sheet not found in {review_path}")
        ws = wb["Review"]

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise SystemExit(f"Empty review sheet in {review_path}")

        missing = REVIEW_REQUIRED_COLUMNS - {h for h in header if h is not None}
        if missing:
            raise SystemExit(
                f"Review sheet missing required columns: {sorted(missing)}"
            )

        col_path = header.index("Path")
        col_type = header.index("Type")
        col_delete = header.index("Delete?")

        marked: list[tuple[str, str]] = []
        for row in rows:
            if not row or row[col_path] is None:
                continue
            v = row[col_delete]
            flag = str(v).strip().upper() if v is not None else ""
            if flag != "Y":
                continue
            kind = "folder" if row[col_type] == "Folder" else "file"
            marked.append((str(row[col_path]), kind))
        return marked
    finally:
        wb.close()


def _filter_redundant_descendants(
    marked: list[tuple[str, str]],
) -> list[tuple[str, str]]:
    marked_paths = {p for p, _ in marked}
    keep = []
    for p, k in marked:
        parts = p.split("/")
        is_descendant = any(
            "/".join(parts[:i]) in marked_paths for i in range(1, len(parts))
        )
        if not is_descendant:
            keep.append((p, k))
    return keep


def _load_terminal_moves(conn) -> set[str]:
    placeholders = ",".join("?" * len(TERMINAL_STATUSES))
    return {
        row[0]
        for row in conn.execute(
            f"SELECT rel_path FROM moves WHERE status IN ({placeholders})",
            TERMINAL_STATUSES,
        )
    }


def _record_move(conn, rel_path, kind, size_bytes, dest_path, status, error):
    with conn:
        conn.execute(
            "INSERT OR REPLACE INTO moves "
            "(rel_path, kind, size_bytes, dest_path, attempted_at, status, error) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (rel_path, kind, size_bytes, dest_path, time.time(), status, error),
        )


def _to_native(rel_posix_path: str) -> str:
    if os.sep != "/":
        return rel_posix_path.replace("/", os.sep)
    return rel_posix_path


def _safe_size(path: str):
    try:
        st = os.stat(path)
    except OSError:
        return None
    if os.path.isdir(path):
        return None
    return st.st_size
